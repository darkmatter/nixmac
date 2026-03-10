import argparse
import csv
import json
import os
import shlex
import shutil
import subprocess
import tempfile
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from git import Actor, Repo
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent.resolve()

# Testcase data
SPREADSHEET: Path = SCRIPT_DIR / "data/dimensions-of-variation.xlsx"

# Location where we store JSON evaluation results during test runs
RESULTS_DIR: Path = SCRIPT_DIR / "data/results"

# Default nix config template
CONFIG_TEMPLATE_DIR: Path = SCRIPT_DIR.parent / "native/templates/nix-darwin-determinate"

# Default place we find nixmac to run evolution during test cases
DEFAULT_NIXMAC = SCRIPT_DIR.parent.parent / "target" / "debug" / "nixmac"

# System nixmac settings.json (~/Library/Application\ Support/com.darkmatter.nixmac/settings.json)
NIXMAC_SETTINGS_PATH = (
    Path.home() / "Library" / "Application Support" / "com.darkmatter.nixmac" / "settings.json"
)

# Populated in __main__ when running as a script
args: argparse.Namespace | None = None


@dataclass
class EvalTestCase:
    row: int
    num: Any
    feature: str
    scenario: str
    persona: str
    priority: str
    request: str
    expected: str
    status: str


def read_test_cases(
    rows: list[int] | None = None, priority: str | None = None, persona: str | None = None
) -> list[EvalTestCase]:
    """Read test cases from the spreadsheet."""

    wb = load_workbook(SPREADSHEET)
    ws = wb["Test Matrix"]
    cases: list[EvalTestCase] = []
    for r in range(5, ws.max_row + 1):
        num = ws.cell(r, 1).value
        if num is None:
            continue
        case = EvalTestCase(
            row=r,
            num=num,
            feature=str(ws.cell(r, 2).value or ""),
            scenario=str(ws.cell(r, 3).value or ""),
            persona=str(ws.cell(r, 4).value or ""),
            priority=str(ws.cell(r, 5).value or ""),
            request=str(ws.cell(r, 6).value or ""),
            expected=str(ws.cell(r, 7).value or ""),
            status=str(ws.cell(r, 8).value or ""),
        )
        if rows and case.num not in rows:
            continue
        if priority and case.priority != priority:
            continue
        if persona and case.persona != persona:
            continue
        if case.request:
            cases.append(case)
    wb.close()
    return cases


def read_test_cases_from_csv(
    csv_path: Path,
    rows: list[int] | None = None,
    priority: str | None = None,
    persona: str | None = None,
) -> list[EvalTestCase]:
    """Read test cases from a CSV file.

    Expected CSV columns:
    - id: test case number
    - prompt: the user request
    - expected_outcome: expected result (succeed/fail_gracefully/refuse)
    - category: high-level category
    - subcategory: more specific scenario
    - quality_dimension: quality aspect being tested
    - notes: additional notes
    """
    cases: list[EvalTestCase] = []

    with open(csv_path, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for idx, row_data in enumerate(reader, start=2):  # Start at 2 (header is row 1)
            try:
                case_id = row_data.get("id", "")
                if not case_id:
                    continue

                # Try to parse id as integer
                num: Any
                try:
                    num = int(case_id)
                except ValueError:
                    num = case_id

                # Map CSV columns to EvalTestCase fields
                case = EvalTestCase(
                    row=idx,
                    num=num,
                    feature=row_data.get("category", ""),
                    scenario=row_data.get("subcategory", ""),
                    persona=row_data.get("quality_dimension", ""),
                    priority="",  # Not in CSV, could be derived from category if needed
                    request=row_data.get("prompt", ""),
                    expected=row_data.get("expected_outcome", ""),
                    status=row_data.get("notes", ""),
                )

                # Apply filters
                if rows and num not in rows:
                    continue
                if priority and case.priority != priority:
                    continue
                if persona and case.persona != persona:
                    continue
                if case.request:
                    cases.append(case)
            except Exception as e:
                print(f"Warning: Skipping row {idx} in CSV due to error: {e}")
                continue

    return cases


def create_nix_config_git_repo():
    """
    Create a temporary nix config git repo from the template, with hostname
    and platform placeholders replaced, ready for nix-darwin evaluation.
    """
    # Create temporary directory
    tmpdir = Path(tempfile.mkdtemp(prefix="nix-config-"))

    # Copy template into temp dir (full recursive copy, preserve all files)
    shutil.copytree(CONFIG_TEMPLATE_DIR, tmpdir, dirs_exist_ok=True)

    # Get the correct macOS LocalHostName for nix-darwin
    try:
        hostname = subprocess.check_output(["scutil", "--get", "LocalHostName"], text=True).strip()
    except subprocess.CalledProcessError:
        # fallback
        hostname = "localhost"

    # Replace placeholders in flake.nix
    flake_path = tmpdir / "flake.nix"
    if flake_path.exists():
        content = flake_path.read_text()
        content = content.replace("HOSTNAME_PLACEHOLDER", hostname)
        content = content.replace("PLATFORM_PLACEHOLDER", "aarch64-darwin")
        flake_path.write_text(content)

    # Initialize git repo
    repo = Repo.init(str(tmpdir))

    # Ensure local user config
    with repo.config_writer() as cw:
        cw.set_value("user", "name", "eval")
        cw.set_value("user", "email", "eval@test")

    # Add all files and commit
    repo.git.add(A=True)
    actor = Actor("eval", "eval@test")
    try:
        repo.index.commit("initial nix config state", author=actor, committer=actor)
    except Exception as exc:
        raise RuntimeError("Failed to create initial git commit for fixture") from exc

    # Refresh index to ensure git tree is clean
    repo.git.update_index("--refresh")

    return tmpdir


def run_test_case(
    case: EvalTestCase,
    nixmac: Path,
    evolve_provider: str | None = None,
    evolve_model: str | None = None,
    summary_provider: str | None = None,
    summary_model: str | None = None,
    openai_key: str | None = None,
    openrouter_key: str | None = None,
    ollama_url: str | None = None,
    host: str | None = None,
) -> Any:
    """Run a single test case.

    Returns provider-specific result (could be string, dict, etc.).
    """
    print(f"Running prompt: {case.request}")

    # Create a git repo in a temporary directory from nix-darwin-determinate
    config_dir: Path | None = None
    result_dir: Path | None = None
    try:
        config_dir = create_nix_config_git_repo()
        print(f"Created git repo with config at: {config_dir}")

        # Generate nixmac settings.json pointing to the config dir
        generate_nixmac_settings(
            evolve_provider,
            evolve_model,
            summary_provider,
            summary_model,
            openai_key,
            openrouter_key,
            ollama_url,
            host,
            str(config_dir),
        )

        # Create a temporary directory to capture the evolution result
        result_dir = Path(tempfile.mkdtemp(prefix="evolution-"))
        out_path = result_dir / "evolution_result.json"

        # Run the test case using nixmac and write output to the temp dir
        cmd = f"{shlex.quote(str(nixmac))} evolve {shlex.quote(case.request)} --out {shlex.quote(str(out_path))}"
        os.system(cmd)

        # Read and return the evolution result if present
        if out_path.exists():
            with open(out_path) as f:
                evolution_result = json.load(f)
            print(json.dumps(evolution_result, indent=2))
            return evolution_result
        else:
            print("No evolution result file found after running nixmac.")
            # TODO: nixmac CLI should report structured failure info (exit code + JSON).
            # For now create a stub JSON result indicating failure so tests can record it.
            stub = {
                "success": False,
                "error": "nixmac did not produce evolution_result.json",
                "case": case.num,
                "command": cmd,
            }
            try:
                with open(out_path, "w") as f:
                    json.dump(stub, f, indent=2)
            except Exception:
                pass
            return stub
    finally:
        if config_dir is not None:
            print(f"Cleaning up temporary config directory: {config_dir}")
            with suppress(Exception):
                shutil.rmtree(config_dir)
        if result_dir is not None:
            print(f"Cleaning up temporary evolution result directory: {result_dir}")
            with suppress(Exception):
                shutil.rmtree(result_dir)


def update_test_case_status(row: int, result: Any) -> None:
    """Persist test result back to the spreadsheet or other storage.

    Not implemented in this script.
    """
    print(f"Writing results JSON to results directory for case row {row}...")
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    result_path = RESULTS_DIR / f"case_{row}_result.json"
    with open(result_path, "w") as f:
        json.dump(result, f, indent=4)
    print(f"Saved results for case {row} to: {result_path}")


def backup_nixmac_settings() -> Path | None:
    """Back up the nixmac settings.json file if it exists, and return the backup path."""
    if NIXMAC_SETTINGS_PATH.exists():
        backup_path = NIXMAC_SETTINGS_PATH.with_suffix(".bak")
        shutil.copy2(NIXMAC_SETTINGS_PATH, backup_path)
        print(f"Backed up nixmac settings to: {backup_path}")
        return backup_path
    else:
        print("No existing nixmac settings found to back up.")
        return None


def restore_nixmac_settings(backup_path: Path | None) -> None:
    """Restore the nixmac settings.json file from the backup path."""
    if backup_path is not None and backup_path.exists():
        shutil.copy2(backup_path, NIXMAC_SETTINGS_PATH)
        backup_path.unlink()
        print(f"Restored nixmac settings from backup: {backup_path}")
    else:
        print("No backup settings found to restore.")


def generate_nixmac_settings(
    evolve_provider: str | None = None,
    evolve_model: str | None = None,
    summary_provider: str | None = None,
    summary_model: str | None = None,
    openai_key: str | None = None,
    openrouter_key: str | None = None,
    ollama_url: str | None = None,
    host: str | None = None,
    configDir: str | None = None,
) -> None:
    """Generate a nixmac settings.json file based on provided parameters."""
    if host is None:
        host = os.uname().nodename

    settings = {
        "evolveProvider": evolve_provider,
        "evolveModel": evolve_model,
        "summaryProvider": summary_provider,
        "summaryModel": summary_model,
        "openaiKey": openai_key,
        "openrouterKey": openrouter_key,
        "ollamaApiBaseUrl": ollama_url,
        "hostAttr": host,
        "configDir": configDir,
    }
    NIXMAC_SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(NIXMAC_SETTINGS_PATH, "w") as f:
        json.dump(settings, f, indent=4)
    print(f"Generated nixmac settings at: {NIXMAC_SETTINGS_PATH}")


def main(parsed_args: argparse.Namespace) -> None:
    # Back up nixmac settings.json before running any test cases, and restore it at the end
    settings_backup_path = backup_nixmac_settings()
    try:
        # Parse comma-delimited rows into list[int]
        rows: list[int] | None
        if parsed_args.rows:
            rows = [int(s) for s in str(parsed_args.rows).replace(",", " ").split() if s]
        else:
            rows = None

        # Read test cases from CSV if specified, otherwise use Excel spreadsheet
        cases: list[EvalTestCase]
        if parsed_args.csv:
            csv_path = Path(parsed_args.csv)
            if not csv_path.exists():
                raise FileNotFoundError(f"CSV file not found: {csv_path}")
            cases = read_test_cases_from_csv(
                csv_path, rows=rows, priority=parsed_args.priority, persona=parsed_args.persona
            )
        else:
            cases = read_test_cases(
                rows=rows, priority=parsed_args.priority, persona=parsed_args.persona
            )
        print(f"Running {len(cases)} test cases...")
        for case in cases:
            print(f"Running case {case.num}: {case.scenario}...")
            nixmac_path = Path(parsed_args.nixmac)
            result = run_test_case(
                case,
                nixmac_path,
                parsed_args.evolve_provider,
                parsed_args.evolve_model,
                parsed_args.summary_provider,
                parsed_args.summary_model,
                parsed_args.openai_key,
                parsed_args.openrouter_key,
                parsed_args.ollama_url,
                parsed_args.host,
            )
            update_test_case_status(case.row, result)
    finally:
        restore_nixmac_settings(settings_backup_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run evaluation test cases.")

    parser.add_argument(
        "--nixmac", type=str, default=str(DEFAULT_NIXMAC), help="Path to nixmac binary"
    )
    # Flags mirroring main nixmac app settings
    parser.add_argument(
        "--evolve-provider", type=str, default="ollama", help="Evolve provider (e.g. openai)"
    )
    parser.add_argument(
        "--evolve-model", type=str, default="gpt-oss-200k:latest", help="Evolve model (e.g. gpt-4)"
    )
    parser.add_argument(
        "--summary-provider", type=str, default="ollama", help="Summary provider (e.g. openai)"
    )
    parser.add_argument(
        "--summary-model", type=str, default="gpt-oss:120b", help="Summary model (e.g. gpt-4)"
    )
    parser.add_argument(
        "--openai-key", dest="openai_key", type=str, default=None, help="OpenAI API key"
    )
    parser.add_argument(
        "--openrouter-key", dest="openrouter_key", type=str, default=None, help="OpenRouter API key"
    )
    parser.add_argument(
        "--ollama-url",
        dest="ollama_url",
        type=str,
        default="",
        help="Ollama base URL (e.g. http://localhost:11434)",
    )
    parser.add_argument("--host", type=str, default=None, help="Host name for your Mac")

    parser.add_argument(
        "--csv",
        type=str,
        default=None,
        help="Path to CSV file containing test prompts (alternative to Excel spreadsheet)",
    )
    parser.add_argument(
        "--rows",
        type=str,
        default=None,
        help="Comma-delimited list of test case numbers to run (e.g., --rows 1,3,5)",
    )
    parser.add_argument(
        "--priority",
        type=str,
        help="Filter test cases by priority (e.g., --priority High), doesn't work for CSV input which doesn't have a priority column",
    )
    parser.add_argument(
        "--persona", type=str, help="Filter test cases by persona (e.g., --persona Developer)"
    )

    args = parser.parse_args()
    main(args)
